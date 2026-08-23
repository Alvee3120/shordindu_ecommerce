"""Excel bulk-import for the catalog: Products / Variations / Addons workbook.

See ProductViewSet.bulk_import / .import_template in views.py for the HTTP layer.
"""

from collections import defaultdict
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .models import (
    Attribute,
    AttributeValue,
    Category,
    Product,
    ProductAddon,
    ProductAttributeValue,
    ProductVariation,
    VariationAttributeValue,
)

REQUIRED_SHEETS = ["Products", "Variations", "Addons"]

PRODUCT_TYPES = {c[0] for c in Product.ProductType.choices}
VISIBILITY_TYPES = {c[0] for c in Product.VisibilityType.choices}
STATUSES = {c[0] for c in Product.Status.choices}

TRUE_VALUES = {"true", "1", "yes", "y"}
FALSE_VALUES = {"false", "0", "no", "n"}


class ImportValidationError(Exception):
    """Workbook-level problem (unreadable file, missing sheet) that aborts the import entirely."""


class RowError(Exception):
    """A single row's data is invalid; the row is skipped and reported, the rest of the import continues."""

    def __init__(self, sheet, row_number, message):
        self.sheet = sheet
        self.row_number = row_number
        self.message = message
        super().__init__(message)


def _clean_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # Excel silently turns a plain-looking code like "1001" into the float 1001.0
        # when a column isn't formatted as text - undo that before it leaks into a SKU.
        return str(int(value))
    return str(value).strip()


def _parse_bool(value, default=False):
    s = _clean_str(value).lower()
    if s == "":
        return default
    if s in TRUE_VALUES:
        return True
    if s in FALSE_VALUES:
        return False
    raise ValueError(f"expected TRUE/FALSE, got {value!r}")


def _parse_decimal(value, field_name, required=True, default=None, min_value=None):
    s = _clean_str(value)
    if s == "":
        if required:
            raise ValueError(f"{field_name} is required")
        return default
    try:
        parsed = Decimal(s)
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} must be a number, got {value!r}") from exc
    if min_value is not None and parsed < min_value:
        raise ValueError(f"{field_name} cannot be negative")
    return parsed


def _parse_int(value, field_name, required=True, default=0, min_value=None):
    s = _clean_str(value)
    if s == "":
        if required:
            raise ValueError(f"{field_name} is required")
        return default
    try:
        parsed = int(float(s))
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a whole number, got {value!r}") from exc
    if min_value is not None and parsed < min_value:
        raise ValueError(f"{field_name} cannot be negative")
    return parsed


def _parse_attributes(value):
    """Parses 'Name:Value;Name2:Value2' into [(name, value), ...]."""
    s = _clean_str(value)
    if not s:
        return []
    pairs = []
    for chunk in s.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        if ":" not in chunk:
            raise ValueError(f"attribute {chunk!r} must be in 'Name:Value' form")
        name, val = chunk.split(":", 1)
        name, val = name.strip(), val.strip()
        if not name or not val:
            raise ValueError(f"attribute {chunk!r} must be in 'Name:Value' form")
        pairs.append((name, val))
    return pairs


def _parse_categories(value):
    """Parses 'Name1;Name2;Name3' into a list of category names."""
    s = _clean_str(value)
    if not s:
        return []
    names = []
    for chunk in s.split(";"):
        chunk = chunk.strip()
        if chunk:
            names.append(chunk)
    return names


def _sheet_rows(ws):
    """Yields (row_number, {header: value}) for a worksheet, skipping the header row and blank rows."""
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    headers = [_clean_str(h) for h in header]
    for row_number, row in enumerate(rows, start=2):
        if all(cell is None or _clean_str(cell) == "" for cell in row):
            continue
        yield row_number, dict(zip(headers, row))


class BulkImporter:
    """One instance per import request - caches resolved categories/attributes for that run only."""

    def __init__(self):
        self._category_cache = {}
        self._attribute_value_cache = {}
        self.errors = []
        self.created = {"products": 0, "variations": 0, "addons": 0}
        self.updated = {"products": 0, "variations": 0, "addons": 0}

    def _err(self, sheet, row_number, message):
        return {"sheet": sheet, "row": row_number, "message": message}

    def resolve_category(self, name):
        key = name.strip().lower()
        if key in self._category_cache:
            return self._category_cache[key]
        category = (
            Category.objects.filter(slug__iexact=name).first()
            or Category.objects.filter(name__iexact=name).first()
        )
        if category is None:
            category = Category.objects.create(name=name.strip())
        self._category_cache[key] = category
        return category

    def resolve_attribute_value(self, attr_name, value):
        key = (attr_name.strip().lower(), value.strip().lower())
        if key in self._attribute_value_cache:
            return self._attribute_value_cache[key]
        attribute = Attribute.objects.filter(name__iexact=attr_name).first()
        if attribute is None:
            attribute = Attribute.objects.create(name=attr_name.strip())
        attribute_value = AttributeValue.objects.filter(attribute=attribute, value__iexact=value).first()
        if attribute_value is None:
            next_order = AttributeValue.objects.filter(attribute=attribute).count()
            attribute_value = AttributeValue.objects.create(
                attribute=attribute, value=value.strip(), sort_order=next_order
            )
        self._attribute_value_cache[key] = attribute_value
        return attribute_value

    def import_workbook(self, file):
        try:
            wb = load_workbook(file, data_only=True, read_only=True)
        except Exception as exc:
            raise ImportValidationError(
                f"Could not read the uploaded file as an Excel (.xlsx) workbook: {exc}"
            ) from exc

        missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
        if missing:
            raise ImportValidationError(f"Workbook is missing required sheet(s): {', '.join(missing)}")

        products_rows = list(_sheet_rows(wb["Products"]))
        variations_rows = list(_sheet_rows(wb["Variations"]))
        addons_rows = list(_sheet_rows(wb["Addons"]))

        variations_by_product = defaultdict(list)
        for row_number, row in variations_rows:
            variations_by_product[_clean_str(row.get("product_sku_prefix"))].append((row_number, row))

        sku_prefix_to_product = {}
        seen_sku_prefixes = set()

        for row_number, row in products_rows:
            sku_prefix = _clean_str(row.get("sku_prefix"))
            if not sku_prefix:
                self.errors.append(self._err("Products", row_number, "sku_prefix is required"))
                continue
            if sku_prefix in seen_sku_prefixes:
                self.errors.append(
                    self._err("Products", row_number, f"duplicate sku_prefix '{sku_prefix}' in this sheet")
                )
                continue
            seen_sku_prefixes.add(sku_prefix)

            try:
                with transaction.atomic():
                    product = self._upsert_product(
                        sku_prefix, row_number, row, variations_by_product.get(sku_prefix, [])
                    )
                sku_prefix_to_product[sku_prefix] = product
            except RowError as exc:
                self.errors.append(self._err(exc.sheet, exc.row_number, exc.message))
            except Exception as exc:  # unexpected DB/constraint errors shouldn't abort the whole import
                self.errors.append(self._err("Products", row_number, f"unexpected error: {exc}"))

        for key, rows in variations_by_product.items():
            if not key:
                for row_number, _ in rows:
                    self.errors.append(self._err("Variations", row_number, "product_sku_prefix is required"))
            elif key not in seen_sku_prefixes:
                for row_number, _ in rows:
                    self.errors.append(
                        self._err("Variations", row_number, f"no Products row with sku_prefix '{key}'")
                    )

        for row_number, row in addons_rows:
            try:
                self._upsert_addon(row_number, row, sku_prefix_to_product)
            except RowError as exc:
                self.errors.append(self._err(exc.sheet, exc.row_number, exc.message))
            except Exception as exc:
                self.errors.append(self._err("Addons", row_number, f"unexpected error: {exc}"))

        return {"created": self.created, "updated": self.updated, "errors": self.errors}

    def _upsert_product(self, sku_prefix, row_number, row, variation_entries):
        name = _clean_str(row.get("name"))
        if not name:
            raise RowError("Products", row_number, "name is required")

        category_names = _parse_categories(row.get("categories"))
        if not category_names:
            raise RowError(
                "Products", row_number, "categories is required (at least one category name, separated by ';')"
            )
        categories = [self.resolve_category(category_name) for category_name in category_names]

        product_type = _clean_str(row.get("product_type")).lower() or Product.ProductType.SIMPLE
        if product_type not in PRODUCT_TYPES:
            raise RowError(
                "Products", row_number, f"product_type must be one of {sorted(PRODUCT_TYPES)}, got {product_type!r}"
            )

        visibility_type = _clean_str(row.get("visibility_type")).lower() or Product.VisibilityType.STANDALONE
        if visibility_type not in VISIBILITY_TYPES:
            raise RowError(
                "Products",
                row_number,
                f"visibility_type must be one of {sorted(VISIBILITY_TYPES)}, got {visibility_type!r}",
            )

        status = _clean_str(row.get("status")).lower() or Product.Status.DRAFT
        if status not in STATUSES:
            raise RowError("Products", row_number, f"status must be one of {sorted(STATUSES)}, got {status!r}")

        description = _clean_str(row.get("description"))

        try:
            product_attr_pairs = _parse_attributes(row.get("attributes"))
        except ValueError as exc:
            raise RowError("Products", row_number, str(exc)) from exc

        if not variation_entries:
            raise RowError(
                "Products",
                row_number,
                "at least one Variations row is required (matched by product_sku_prefix)",
            )

        parsed_variations = []
        seen_variation_skus = set()
        for v_row_number, v_row in variation_entries:
            try:
                variation_sku = _clean_str(v_row.get("variation_sku")) or f"{sku_prefix}-DEFAULT"
                if variation_sku in seen_variation_skus:
                    raise ValueError(f"duplicate variation_sku '{variation_sku}' for this product")
                seen_variation_skus.add(variation_sku)
                other_owner = (
                    ProductVariation.objects.filter(sku=variation_sku)
                    .exclude(product__sku_prefix=sku_prefix)
                    .exists()
                )
                if other_owner:
                    raise ValueError(f"variation_sku '{variation_sku}' is already used by another product")
                price = _parse_decimal(v_row.get("price"), "price", min_value=0)
                compare_at_price = _parse_decimal(
                    v_row.get("compare_at_price"), "compare_at_price", required=False, min_value=0
                )
                stock_quantity = _parse_int(v_row.get("stock_quantity"), "stock_quantity", min_value=0)
                is_active = _parse_bool(v_row.get("is_active"), default=True)
                attribute_pairs = _parse_attributes(v_row.get("attributes"))
            except ValueError as exc:
                raise RowError("Variations", v_row_number, str(exc)) from exc

            parsed_variations.append(
                {
                    "sku": variation_sku,
                    "price": price,
                    "compare_at_price": compare_at_price,
                    "stock_quantity": stock_quantity,
                    "is_active": is_active,
                    "attribute_pairs": attribute_pairs,
                }
            )

        existing = Product.objects.filter(sku_prefix=sku_prefix).first()
        is_new = existing is None
        product = existing or Product(sku_prefix=sku_prefix)
        product.name = name
        product.description = description
        product.product_type = product_type
        product.visibility_type = visibility_type
        product.status = status
        product.save()

        product.categories.set(categories)

        product.product_attribute_values.all().delete()
        for attr_name, value in product_attr_pairs:
            attribute_value = self.resolve_attribute_value(attr_name, value)
            ProductAttributeValue.objects.create(product=product, attribute_value=attribute_value)

        for parsed in parsed_variations:
            variation_existing = ProductVariation.objects.filter(sku=parsed["sku"]).first()
            variation_is_new = variation_existing is None
            variation = variation_existing or ProductVariation(sku=parsed["sku"])
            variation.product = product
            variation.price = parsed["price"]
            variation.compare_at_price = parsed["compare_at_price"]
            variation.stock_quantity = parsed["stock_quantity"]
            variation.is_active = parsed["is_active"]
            variation.save()

            variation.variation_attribute_values.all().delete()
            for attr_name, value in parsed["attribute_pairs"]:
                attribute_value = self.resolve_attribute_value(attr_name, value)
                VariationAttributeValue.objects.create(variation=variation, attribute_value=attribute_value)

            if variation_is_new:
                self.created["variations"] += 1
            else:
                self.updated["variations"] += 1

        if is_new:
            self.created["products"] += 1
        else:
            self.updated["products"] += 1

        return product

    def _upsert_addon(self, row_number, row, sku_prefix_to_product):
        parent_sku = _clean_str(row.get("parent_sku_prefix"))
        addon_sku = _clean_str(row.get("addon_sku_prefix"))
        if not parent_sku or not addon_sku:
            raise RowError("Addons", row_number, "parent_sku_prefix and addon_sku_prefix are required")

        parent = sku_prefix_to_product.get(parent_sku) or Product.objects.filter(sku_prefix=parent_sku).first()
        addon = sku_prefix_to_product.get(addon_sku) or Product.objects.filter(sku_prefix=addon_sku).first()
        if parent is None:
            raise RowError("Addons", row_number, f"no product with sku_prefix '{parent_sku}'")
        if addon is None:
            raise RowError("Addons", row_number, f"no product with sku_prefix '{addon_sku}'")
        if parent.pk == addon.pk:
            raise RowError("Addons", row_number, "a product cannot be its own addon")

        try:
            is_required = _parse_bool(row.get("is_required"), default=False)
            min_select = _parse_int(row.get("min_select"), "min_select", required=False, default=0, min_value=0)
            max_select = _parse_int(row.get("max_select"), "max_select", required=False, default=1, min_value=0)
            price_override = _parse_decimal(
                row.get("price_override"), "price_override", required=False, min_value=0
            )
            sort_order = _parse_int(row.get("sort_order"), "sort_order", required=False, default=0, min_value=0)
        except ValueError as exc:
            raise RowError("Addons", row_number, str(exc)) from exc

        _, created = ProductAddon.objects.update_or_create(
            parent_product=parent,
            addon_product=addon,
            defaults={
                "is_required": is_required,
                "min_select": min_select,
                "max_select": max_select,
                "price_override": price_override,
                "sort_order": sort_order,
            },
        )
        if created:
            self.created["addons"] += 1
        else:
            self.updated["addons"] += 1


def build_template_workbook():
    """Builds the downloadable .xlsx template: Products/Variations/Addons sheets, instructions, and a live
    reference list of existing categories."""
    wb = Workbook()

    products_ws = wb.active
    products_ws.title = "Products"
    products_ws.append(
        ["sku_prefix", "name", "categories", "description", "product_type", "visibility_type", "status", "attributes"]
    )
    products_ws.append(
        [
            "SHIRT-001",
            "Classic Cotton Shirt",
            "Men's Clothing;New Arrivals",
            "A comfortable everyday shirt",
            "variable",
            "standalone",
            "active",
            "Brand:Acme",
        ]
    )

    variations_ws = wb.create_sheet("Variations")
    variations_ws.append(
        ["product_sku_prefix", "variation_sku", "attributes", "price", "compare_at_price", "stock_quantity", "is_active"]
    )
    variations_ws.append(["SHIRT-001", "SHIRT-001-S-BLU", "Size:S;Color:Blue", "19.99", "24.99", "50", "TRUE"])
    variations_ws.append(["SHIRT-001", "SHIRT-001-M-BLU", "Size:M;Color:Blue", "19.99", "24.99", "40", "TRUE"])

    addons_ws = wb.create_sheet("Addons")
    addons_ws.append(
        ["parent_sku_prefix", "addon_sku_prefix", "is_required", "min_select", "max_select", "price_override", "sort_order"]
    )
    addons_ws.append(["SHIRT-001", "TIE-001", "FALSE", "0", "1", "", "0"])

    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        (
            "How this workbook works",
            "3 data sheets - Products, Variations, Addons. Every product needs one row in Products and at "
            "least one row in Variations, linked by sku_prefix / product_sku_prefix. Re-uploading this file "
            "updates existing rows that match the same sku_prefix / variation_sku instead of duplicating them.",
        ),
        ("", ""),
        ("Products sheet", ""),
        ("sku_prefix", "Required, unique. Your own short product code, e.g. SHIRT-001. Links rows in the other sheets."),
        ("name", "Required."),
        (
            "categories",
            "Required, at least one. Semicolon-separated list of existing category names/slugs (see the "
            "Categories sheet) or new ones to create, e.g. Men's Clothing;New Arrivals.",
        ),
        ("description", "Optional."),
        ("product_type", "simple or variable. Default: simple."),
        ("visibility_type", "standalone, addon_only, or both. Default: standalone."),
        ("status", "draft, active, or archived. Default: draft."),
        ("attributes", "Optional, non-variant attributes, e.g. Material:Cotton;Brand:Acme."),
        ("", ""),
        ("Variations sheet", ""),
        ("product_sku_prefix", "Required. Must match a sku_prefix from the Products sheet."),
        ("variation_sku", "Optional. If blank, generated as {product_sku_prefix}-DEFAULT."),
        ("attributes", "e.g. Size:M;Color:Blue. Leave blank for simple products."),
        ("price", "Required."),
        ("compare_at_price", "Optional 'was' price."),
        ("stock_quantity", "Required."),
        ("is_active", "TRUE or FALSE. Default: TRUE."),
        ("", ""),
        ("Addons sheet", "One row per parent product -> addon product relationship."),
        ("parent_sku_prefix", "Required. The main product."),
        ("addon_sku_prefix", "Required. The product offered as an addon to the parent."),
        ("is_required", "TRUE or FALSE. Default: FALSE."),
        ("min_select / max_select", "How many of this addon may be chosen. Defaults: 0 / 1."),
        ("price_override", "Optional price for the addon, overriding its own price."),
        ("sort_order", "Optional display order. Default: 0."),
        ("", ""),
        ("Images", "Not included here - add product photos from the admin Products page after importing."),
    ]
    for row in instructions:
        instructions_ws.append(row)
    instructions_ws.column_dimensions["A"].width = 28
    instructions_ws.column_dimensions["B"].width = 95
    for cell in instructions_ws["A"]:
        cell.font = Font(bold=True)

    categories_ws = wb.create_sheet("Categories")
    categories_ws.append(["name", "slug"])
    for category in Category.objects.order_by("name"):
        categories_ws.append([category.name, category.slug])

    for ws in (products_ws, variations_ws, addons_ws, categories_ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    return wb

import io

from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from rest_framework.test import APITestCase

from apps.catalog.models import Category, Product, ProductAddon, ProductVariation
from apps.users.models import User

PRODUCTS_HEADER = [
    "sku_prefix",
    "name",
    "categories",
    "description",
    "product_type",
    "visibility_type",
    "status",
    "attributes",
]
VARIATIONS_HEADER = [
    "product_sku_prefix",
    "variation_sku",
    "attributes",
    "price",
    "compare_at_price",
    "stock_quantity",
    "is_active",
]
ADDONS_HEADER = [
    "parent_sku_prefix",
    "addon_sku_prefix",
    "is_required",
    "min_select",
    "max_select",
    "price_override",
    "sort_order",
]


def build_upload(products=None, variations=None, addons=None):
    wb = Workbook()
    products_ws = wb.active
    products_ws.title = "Products"
    products_ws.append(PRODUCTS_HEADER)
    for row in products or []:
        products_ws.append(row)

    variations_ws = wb.create_sheet("Variations")
    variations_ws.append(VARIATIONS_HEADER)
    for row in variations or []:
        variations_ws.append(row)

    addons_ws = wb.create_sheet("Addons")
    addons_ws.append(ADDONS_HEADER)
    for row in addons or []:
        addons_ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        "import.xlsx",
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class BulkImportTests(APITestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            email="staff@example.com", password="pw", name="Staff", is_staff=True
        )
        self.customer = User.objects.create_user(
            email="customer@example.com", password="pw", name="Customer"
        )
        self.category = Category.objects.create(name="Shirts")

    def test_creates_products_variations_and_addon_link(self):
        upload = build_upload(
            products=[
                ["SHIRT-001", "Classic Shirt", "Shirts", "", "variable", "standalone", "active", ""],
                ["TIE-001", "Silk Tie", "Shirts", "", "simple", "addon_only", "active", ""],
            ],
            variations=[
                ["SHIRT-001", "SHIRT-001-S", "Size:S", "19.99", "", "10", "TRUE"],
                ["SHIRT-001", "SHIRT-001-M", "Size:M", "19.99", "", "5", "TRUE"],
                ["TIE-001", "", "", "9.99", "", "20", "TRUE"],
            ],
            addons=[["SHIRT-001", "TIE-001", "FALSE", "0", "1", "", "0"]],
        )
        self.client.force_authenticate(self.staff)
        response = self.client.post("/api/products/bulk-import/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["created"], {"products": 2, "variations": 3, "addons": 1})
        self.assertEqual(response.data["errors"], [])

        shirt = Product.objects.get(sku_prefix="SHIRT-001")
        self.assertEqual(shirt.variations.count(), 2)
        tie = Product.objects.get(sku_prefix="TIE-001")
        self.assertEqual(tie.variations.get().sku, "TIE-001-DEFAULT")
        self.assertTrue(ProductAddon.objects.filter(parent_product=shirt, addon_product=tie).exists())

    def test_reimport_updates_existing_records_instead_of_duplicating(self):
        first = build_upload(
            products=[["SHIRT-001", "Classic Shirt", "Shirts", "", "simple", "standalone", "draft", ""]],
            variations=[["SHIRT-001", "SHIRT-001-DEFAULT", "", "19.99", "", "10", "TRUE"]],
        )
        self.client.force_authenticate(self.staff)
        self.client.post("/api/products/bulk-import/", {"file": first}, format="multipart")

        second = build_upload(
            products=[["SHIRT-001", "Classic Shirt V2", "Shirts", "", "simple", "standalone", "active", ""]],
            variations=[["SHIRT-001", "SHIRT-001-DEFAULT", "", "24.99", "", "8", "TRUE"]],
        )
        response = self.client.post("/api/products/bulk-import/", {"file": second}, format="multipart")

        self.assertEqual(response.data["created"], {"products": 0, "variations": 0, "addons": 0})
        self.assertEqual(response.data["updated"]["products"], 1)
        self.assertEqual(response.data["updated"]["variations"], 1)
        self.assertEqual(Product.objects.filter(sku_prefix="SHIRT-001").count(), 1)
        product = Product.objects.get(sku_prefix="SHIRT-001")
        self.assertEqual(product.name, "Classic Shirt V2")
        self.assertEqual(product.status, "active")
        variation = ProductVariation.objects.get(sku="SHIRT-001-DEFAULT")
        self.assertEqual(str(variation.price), "24.99")

    def test_auto_creates_unknown_category(self):
        upload = build_upload(
            products=[["HAT-001", "Sun Hat", "Headwear", "", "simple", "standalone", "active", ""]],
            variations=[["HAT-001", "", "", "12.00", "", "3", "TRUE"]],
        )
        self.client.force_authenticate(self.staff)
        response = self.client.post("/api/products/bulk-import/", {"file": upload}, format="multipart")

        self.assertEqual(response.data["created"]["products"], 1)
        self.assertTrue(Category.objects.filter(name="Headwear").exists())

    def test_creates_product_with_multiple_categories(self):
        upload = build_upload(
            products=[
                ["MULTI-001", "Multi Cat Product", "Shirts;Headwear", "", "simple", "standalone", "active", ""],
            ],
            variations=[["MULTI-001", "", "", "15.00", "", "5", "TRUE"]],
        )
        self.client.force_authenticate(self.staff)
        response = self.client.post("/api/products/bulk-import/", {"file": upload}, format="multipart")

        self.assertEqual(response.data["created"]["products"], 1, response.data)
        product = Product.objects.get(sku_prefix="MULTI-001")
        self.assertEqual(set(product.categories.values_list("name", flat=True)), {"Shirts", "Headwear"})

    def test_missing_categories_is_reported(self):
        upload = build_upload(
            products=[["NOCAT-001", "No Category", "", "", "simple", "standalone", "active", ""]],
            variations=[["NOCAT-001", "", "", "10.00", "", "1", "TRUE"]],
        )
        self.client.force_authenticate(self.staff)
        response = self.client.post("/api/products/bulk-import/", {"file": upload}, format="multipart")

        self.assertEqual(response.data["created"]["products"], 0)
        self.assertFalse(Product.objects.filter(sku_prefix="NOCAT-001").exists())
        self.assertEqual(len(response.data["errors"]), 1)
        self.assertIn("categories", response.data["errors"][0]["message"])

    def test_invalid_row_is_reported_without_blocking_valid_rows(self):
        upload = build_upload(
            products=[
                ["GOOD-001", "Good Product", "Shirts", "", "simple", "standalone", "active", ""],
                ["BAD-001", "Bad Product", "Shirts", "", "not-a-type", "standalone", "active", ""],
            ],
            variations=[
                ["GOOD-001", "", "", "10.00", "", "1", "TRUE"],
                ["BAD-001", "", "", "10.00", "", "1", "TRUE"],
            ],
        )
        self.client.force_authenticate(self.staff)
        response = self.client.post("/api/products/bulk-import/", {"file": upload}, format="multipart")

        self.assertEqual(response.data["created"]["products"], 1)
        self.assertTrue(Product.objects.filter(sku_prefix="GOOD-001").exists())
        self.assertFalse(Product.objects.filter(sku_prefix="BAD-001").exists())
        self.assertEqual(len(response.data["errors"]), 1)
        self.assertEqual(response.data["errors"][0]["sheet"], "Products")

    def test_requires_staff(self):
        upload = build_upload(
            products=[["X-001", "X", "Shirts", "", "simple", "standalone", "active", ""]],
            variations=[["X-001", "", "", "1.00", "", "1", "TRUE"]],
        )
        self.client.force_authenticate(self.customer)
        response = self.client.post("/api/products/bulk-import/", {"file": upload}, format="multipart")
        self.assertEqual(response.status_code, 403)

    def test_import_template_download(self):
        self.client.force_authenticate(self.staff)
        response = self.client.get("/api/products/import-template/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class ProductCategoryAPITests(APITestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            email="staff2@example.com", password="pw", name="Staff", is_staff=True
        )
        self.shirts = Category.objects.create(name="Shirts")
        self.headwear = Category.objects.create(name="Headwear")
        self.client.force_authenticate(self.staff)

    def test_create_product_with_multiple_categories(self):
        response = self.client.post(
            "/api/products/",
            {
                "name": "Bucket Hat",
                "categories": [self.shirts.id, self.headwear.id],
                "sku_prefix": "BUCKET-001",
                "variations": [{"sku": "BUCKET-001-DEFAULT", "price": "20.00", "stock_quantity": 5}],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(set(response.data["categories"]), {self.shirts.id, self.headwear.id})
        self.assertEqual(set(response.data["category_names"]), {"Shirts", "Headwear"})

    def test_empty_categories_is_rejected(self):
        response = self.client.post(
            "/api/products/",
            {
                "name": "No Category Product",
                "categories": [],
                "sku_prefix": "NOCAT-002",
                "variations": [{"sku": "NOCAT-002-DEFAULT", "price": "20.00", "stock_quantity": 5}],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("categories", response.data)

    def test_patch_replaces_categories(self):
        product = Product.objects.create(name="Cap", sku_prefix="CAP-001")
        product.categories.set([self.shirts.id])
        ProductVariation.objects.create(product=product, sku="CAP-001-DEFAULT", price="10.00", stock_quantity=1)

        response = self.client.patch(
            f"/api/products/{product.id}/", {"categories": [self.headwear.id]}, format="json"
        )
        self.assertEqual(response.status_code, 200, response.data)
        product.refresh_from_db()
        self.assertEqual(list(product.categories.values_list("id", flat=True)), [self.headwear.id])

    def test_filter_by_multiple_category_ids_matches_any(self):
        p1 = Product.objects.create(name="Shirt A", sku_prefix="A-001")
        p1.categories.set([self.shirts.id])
        ProductVariation.objects.create(product=p1, sku="A-001-DEFAULT", price="10.00", stock_quantity=1)

        p2 = Product.objects.create(name="Hat B", sku_prefix="B-001")
        p2.categories.set([self.headwear.id])
        ProductVariation.objects.create(product=p2, sku="B-001-DEFAULT", price="10.00", stock_quantity=1)

        response = self.client.get(f"/api/products/?category={self.shirts.id},{self.headwear.id}")
        self.assertEqual(response.status_code, 200)
        names = {p["name"] for p in response.data["results"]}
        self.assertEqual(names, {"Shirt A", "Hat B"})
